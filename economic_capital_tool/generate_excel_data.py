"""
Generate the dummy bank portfolio data as a formatted Excel workbook.
Produces 1 year of data across all product types with proper financial formatting.
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from data.generator import generate_all_data, RATINGS, MIGRATION_MATRIX, GICS_GROUP_DISPLAY_NAMES

OUTPUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'Bank_Portfolio_Dummy_Data.xlsx')

HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor='1a1a2e')
DATA_FONT = Font(name='Arial', size=10)
BLUE_FONT = Font(name='Arial', size=10, color='0000FF')
BORDER = Border(
    bottom=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC')
)
SECTION_FILL = PatternFill('solid', fgColor='e8eaf6')
SECTION_FONT = Font(name='Arial', bold=True, size=12, color='1a1a2e')
PCT_FMT = '0.00%'
NUM_FMT = '#,##0'
CURR_FMT = '$#,##0;($#,##0);"-"'
CURR_DEC_FMT = '$#,##0.00;($#,##0.00);"-"'
BPS_FMT = '#,##0.0'


def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def auto_width(ws, max_col, max_width=30):
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        max_len = 0
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def write_counterparties(wb, counterparties):
    ws = wb.create_sheet('Counterparties')
    headers = ['Counterparty ID', 'Legal Name', 'Parent ID', 'GICS Industry Group',
               'GICS Sector', 'Country', 'Region', 'Rating', 'PD (1Y)', 'RSQ',
               'Public', 'Revenue ($mm)', 'Total Assets ($mm)']
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for cp in counterparties:
        row = [cp['counterparty_id'], cp['legal_name'], cp.get('parent_id', ''),
               GICS_GROUP_DISPLAY_NAMES.get(cp['sector_code'], cp['sector_code']),
               cp.get('gics_sector', ''),
               cp['country_code'], cp['region_code'],
               cp['rating'], cp['pd_1y'], cp['rsq'],
               'Yes' if cp['is_public'] else 'No',
               cp['revenue_mm'], cp['total_assets_mm']]
        ws.append(row)

    for r in range(2, ws.max_row + 1):
        ws.cell(r, 9).number_format = PCT_FMT
        ws.cell(r, 9).font = BLUE_FONT
        ws.cell(r, 10).number_format = '0.0000'
        ws.cell(r, 10).font = BLUE_FONT
        ws.cell(r, 12).number_format = CURR_FMT
        ws.cell(r, 13).number_format = CURR_FMT
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).border = BORDER

    auto_width(ws, len(headers))
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{ws.max_row}'


def write_instruments(wb, instruments):
    ws = wb.create_sheet('Instruments')
    headers = ['Instrument ID', 'Type', 'Counterparty ID', 'Currency', 'Maturity Date',
               'Seniority', 'LGD', 'Rating', 'Drawn Amount', 'Undrawn Amount',
               'Notional', 'MTM Value', 'Interest Rate', 'Collateral Type',
               'CCF', 'Subtype/Direction', 'CDS Spread (bps)']
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for inst in instruments:
        subtype = inst.get('derivative_subtype', inst.get('cds_direction',
                  inst.get('trade_type', inst.get('guarantee_type', ''))))
        row = [
            inst['instrument_id'], inst['instrument_type'], inst['counterparty_id'],
            inst['currency'], inst['maturity_date'], inst['seniority'],
            inst['lgd'], inst.get('rating', ''),
            inst.get('drawn_amount', 0), inst.get('undrawn_amount', 0),
            inst.get('notional', 0), inst.get('mtm_value', 0),
            inst.get('interest_rate', ''), inst.get('collateral_type', ''),
            inst.get('ccf', ''), subtype, inst.get('cds_spread_bps', '')
        ]
        ws.append(row)

    for r in range(2, ws.max_row + 1):
        ws.cell(r, 7).number_format = PCT_FMT
        ws.cell(r, 7).font = BLUE_FONT
        for c in [9, 10, 11, 12]:
            ws.cell(r, c).number_format = CURR_DEC_FMT
        if ws.cell(r, 13).value and ws.cell(r, 13).value != '':
            ws.cell(r, 13).number_format = PCT_FMT
            ws.cell(r, 13).font = BLUE_FONT
        if ws.cell(r, 15).value and ws.cell(r, 15).value != '':
            ws.cell(r, 15).number_format = PCT_FMT
            ws.cell(r, 15).font = BLUE_FONT
        if ws.cell(r, 17).value and ws.cell(r, 17).value != '':
            ws.cell(r, 17).number_format = BPS_FMT
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).border = BORDER

    auto_width(ws, len(headers))
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{ws.max_row}'


def write_snapshots(wb, snapshots):
    ws = wb.create_sheet('Monthly Snapshots')
    headers = ['Snapshot Date', 'Instrument ID', 'Counterparty ID', 'Type',
               'Drawn Amount', 'Undrawn Amount', 'Notional', 'MTM Value', 'EAD']
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    # Write only a sample (first 3 months) to keep file manageable
    sample = [s for s in snapshots if s['snapshot_date'] <= '2025-06-01']
    for snap in sample:
        ws.append([snap['snapshot_date'], snap['instrument_id'], snap['counterparty_id'],
                   snap['instrument_type'], snap['drawn_amount'], snap['undrawn_amount'],
                   snap['notional'], snap['mtm_value'], snap['ead']])

    for r in range(2, ws.max_row + 1):
        for c in [5, 6, 7, 8, 9]:
            ws.cell(r, c).number_format = CURR_DEC_FMT
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).border = BORDER

    auto_width(ws, len(headers))
    ws.freeze_panes = 'A2'

    # Add note about full data
    note_row = ws.max_row + 2
    ws.cell(note_row, 1).value = "Note: Full dataset contains 12 monthly snapshots. Showing first 3 months here."
    ws.cell(note_row, 1).font = Font(name='Arial', italic=True, color='666666')


def write_pd_curves(wb, pd_curves):
    ws = wb.create_sheet('PD Curves')
    headers = ['Counterparty ID', 'Curve Type', 'Tenor (Years)', 'PD', 'Source', 'Model Version']
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    # Sample: first 200 counterparties
    cp_ids = sorted(set(pc['counterparty_id'] for pc in pd_curves))[:200]
    for pc in pd_curves:
        if pc['counterparty_id'] in cp_ids:
            ws.append([pc['counterparty_id'], pc['curve_type'], pc['tenor'],
                       pc['pd'], pc['source'], pc['model_version']])

    for r in range(2, ws.max_row + 1):
        ws.cell(r, 4).number_format = '0.000000'
        ws.cell(r, 4).font = BLUE_FONT
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).border = BORDER

    auto_width(ws, len(headers))
    ws.freeze_panes = 'A2'


def write_migration_matrix(wb):
    ws = wb.create_sheet('Migration Matrix')
    ws.cell(1, 1).value = '1-Year Rating Migration Probabilities'
    ws.cell(1, 1).font = SECTION_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(RATINGS)+1)

    ws.cell(3, 1).value = 'From \\ To'
    ws.cell(3, 1).font = HEADER_FONT
    ws.cell(3, 1).fill = HEADER_FILL
    for j, rating in enumerate(RATINGS):
        ws.cell(3, j+2).value = rating
        ws.cell(3, j+2).font = HEADER_FONT
        ws.cell(3, j+2).fill = HEADER_FILL
        ws.cell(3, j+2).alignment = Alignment(horizontal='center')

    for i, rating in enumerate(RATINGS):
        ws.cell(i+4, 1).value = rating
        ws.cell(i+4, 1).font = Font(name='Arial', bold=True, size=10)
        ws.cell(i+4, 1).fill = PatternFill('solid', fgColor='d1d5db')
        for j in range(len(RATINGS)):
            cell = ws.cell(i+4, j+2)
            cell.value = MIGRATION_MATRIX[i, j]
            cell.number_format = '0.0000%'
            cell.alignment = Alignment(horizontal='center')
            cell.border = BORDER
            if i == j:
                cell.fill = PatternFill('solid', fgColor='d4edda')
            elif j == len(RATINGS) - 1 and MIGRATION_MATRIX[i, j] > 0.001:
                cell.fill = PatternFill('solid', fgColor='f8d7da')
                cell.font = Font(name='Arial', size=10, color='FF0000')

    auto_width(ws, len(RATINGS) + 1)


def write_scenarios(wb, scenarios):
    ws = wb.create_sheet('Scenarios')
    headers = ['Scenario ID', 'Scenario Name', 'Type', 'Description',
               'GDP Shock (σ)', 'Rate Shock (σ)', 'Spread Shock (σ)', 'Factor Shocks']
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for sc in scenarios:
        shocks_str = '; '.join(f"{k}: {v:+.1f}σ" for k, v in sc['factor_shocks'].items()) if sc['factor_shocks'] else 'None'
        ws.append([sc['scenario_id'], sc['scenario_name'], sc['scenario_type'],
                   sc['description'], sc['gdp_shock'], sc['rate_shock'],
                   sc['spread_shock'], shocks_str])

    for r in range(2, ws.max_row + 1):
        for c in [5, 6, 7]:
            cell = ws.cell(r, c)
            cell.font = BLUE_FONT
            if cell.value and float(cell.value) < 0:
                cell.font = Font(name='Arial', size=10, color='FF0000')
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).border = BORDER

    auto_width(ws, len(headers), max_width=50)
    ws.freeze_panes = 'A2'


def write_summary(wb, data):
    ws = wb.create_sheet('Summary')
    ws.sheet_properties.tabColor = '1a1a2e'

    ws.cell(1, 1).value = 'Corporate Bank Portfolio - Dummy Data Summary'
    ws.cell(1, 1).font = Font(name='Arial', bold=True, size=16, color='1a1a2e')
    ws.merge_cells('A1:F1')

    ws.cell(2, 1).value = 'Generated for Economic Capital Simulation Tool'
    ws.cell(2, 1).font = Font(name='Arial', italic=True, size=11, color='666666')

    info = [
        ('Portfolio Statistics', '', True),
        ('Number of Counterparties', len(data['counterparties']), False),
        ('Number of Instruments', len(data['instruments']), False),
        ('Monthly Snapshots', '12 (Mar 2025 - Feb 2026)', False),
        ('', '', False),
        ('Product Mix', '', True),
        ('Term Loans', sum(1 for i in data['instruments'] if i['instrument_type']=='TermLoan'), False),
        ('Revolving Credit', sum(1 for i in data['instruments'] if i['instrument_type']=='Revolver'), False),
        ('IR Derivatives', sum(1 for i in data['instruments'] if i['instrument_type']=='Derivative_IR'), False),
        ('FX Derivatives', sum(1 for i in data['instruments'] if i['instrument_type']=='Derivative_FX'), False),
        ('Credit Default Swaps', sum(1 for i in data['instruments'] if i['instrument_type']=='CDS'), False),
        ('Trade Finance', sum(1 for i in data['instruments'] if i['instrument_type']=='TradeFinance'), False),
        ('Guarantees', sum(1 for i in data['instruments'] if i['instrument_type']=='Guarantee'), False),
        ('', '', False),
        ('Geographic Coverage', '', True),
        ('Countries', len(set(cp['country_code'] for cp in data['counterparties'])), False),
        ('GICS Industry Groups', len(set(cp['sector_code'] for cp in data['counterparties'])), False),
        ('GICS Sectors', len(set(cp.get('gics_sector', '') for cp in data['counterparties'])), False),
        ('', '', False),
        ('Data Sheets', '', True),
        ('Counterparties', 'Master counterparty data with ratings, PDs, and factor loadings', False),
        ('Instruments', 'Full instrument detail by product type', False),
        ('Monthly Snapshots', 'Time series exposure data (sample: 3 months)', False),
        ('PD Curves', 'PD term structures by counterparty (sample: 200 CPs)', False),
        ('Migration Matrix', '8-state annual transition probabilities', False),
        ('Scenarios', '5 pre-defined stress scenarios with factor shocks', False),
    ]

    for idx, (label, value, is_section) in enumerate(info):
        r = idx + 4
        if is_section:
            ws.cell(r, 1).value = label
            ws.cell(r, 1).font = Font(name='Arial', bold=True, size=12, color='1a1a2e')
            ws.cell(r, 1).fill = SECTION_FILL
            ws.cell(r, 2).fill = SECTION_FILL
        elif label:
            ws.cell(r, 1).value = label
            ws.cell(r, 1).font = DATA_FONT
            ws.cell(r, 2).value = value
            ws.cell(r, 2).font = Font(name='Arial', size=10, bold=True)
            ws.cell(r, 1).border = BORDER
            ws.cell(r, 2).border = BORDER

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 55


def main():
    print("Generating portfolio data...")
    data = generate_all_data(n_counterparties=500, n_instruments=2000, seed=42)

    print("Creating Excel workbook...")
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    write_summary(wb, data)
    write_counterparties(wb, data['counterparties'])
    write_instruments(wb, data['instruments'])
    write_snapshots(wb, data['snapshots'])
    write_pd_curves(wb, data['pd_curves'])
    write_migration_matrix(wb)
    write_scenarios(wb, data['scenarios'])

    # Set Summary as active
    wb.active = 0

    print(f"Saving to {OUTPUT_PATH}...")
    wb.save(OUTPUT_PATH)
    print(f"Done! File saved to: {OUTPUT_PATH}")
    return OUTPUT_PATH


if __name__ == '__main__':
    main()
